
from io import BytesIO
from time import sleep
import requests
from urllib.parse import unquote, urlencode, urlparse, urljoin
import genshinstats as gs
from json import load, dump
from os import getcwd, remove
from datetime import datetime
import openpyxl
from os.path import exists
from nextcord import Embed, Member
from prettytable import PrettyTable as pt


def logt(*arg):
    txt = ' '.join([str(i) for i in arg])
    with open('log.txt', 'a+') as f:
        f.write(txt) 

class WishClient:
    


    def __init__(self, bot):
        self.banners = {}
        self.bot = bot
        self.resm = bot.resource_manager
        self.wish_rawfolder = self.resm.db.format(path='wishes/raw/')
        self.wish_processedfolder = self.resm.db.format(path='wishes/processed/')
        
        self.wish_processedfolder = getcwd()+'/assets/wishes/processed/'
        self.wish_rawfolder = getcwd()+'/assets/wishes/raw/'
        #with open(getcwd()+'/assets/data/banners.json', 'r') as f:
            #self.banners = load(f)
        with open(self.resm.path.format(path='data/banners.json'), 'r') as f:
            self.banners = load(f)
        self.URL_MAIN = 'https://hk4e-api-os.hoyoverse.com/event/gacha_info/api/getGachaLog?authkey_ver=1&gacha_id={gacha_id}&device_type={device_type}&sign_type=2&auth_appid=webview_gacha&init_type=301&lang=en&region={region}&authkey={authkey}&gacha_type={type}&page={page}&size={size}&end_id={end_id}&game_biz=hk4e_global'

        self.BANNER_MAP = {
            '301': 'characters',
            '302' : 'weapons',
            '400': 'characters',
            '100' : 'novice',
            '200' : 'standard'
        }


        self.BANNER_SECOND_MAP = {
            'wanderlust': '200',
            'beginner': '100',
            'epitome': '302'
        }
    


    def generate_query(self, url: str, **kwargs):
        '''

        generates url for wish history request from powershell url
        '''

        MAIN_URL = 'https://hk4e-api-os.hoyoverse.com/event/gacha_info/api/getGachaLog?'
        query_raw = urlparse(url).query
        query_dict = {

        }

        for q in query_raw.split('&'):
            if q.split('=')[0] not in ['gacha_type', 'page', 'size', 'end_id']:
                query_dict[q.split('=')[0].strip()] = unquote(q.split("=")[1].strip())
        for k in kwargs:
            query_dict[k] = kwargs[k]
        return MAIN_URL+ urlencode(query_dict)







    def get_wish_history(self, authkey_url: str, **kwargs):
        '''

        gets wish history

        kwargs
        ---------
        gacha_type: 301, 302, 200, 100
        page : 1....
        size: 6 or 20
        end_id: last id of the pull of previous page or just zero

        '''

        url = self.generate_query(authkey_url, **kwargs)   
        if url is not None:     
            data =  requests.get(url).json()
            if 'data' in data:
                if 'list' in data['data']:
                    return data['data']['list'] if len(data['data']['list']) > 0 else None


    def get_full_wishhistory(self, authkey_url:str, **kwargs):
        '''

        fetches all history of all banners even if no kwargs are provided

        '''
        full_wish = []

        for ban in [301, 302, 200, 100]:
            kwargs['gacha_type'] = ban
            kwargs['end_id'] = 0
            kwargs['page'] = 1
            kwargs['size'] = 20
            data = self.get_wish_history(authkey_url, **kwargs)
            while data is not None:
                kwargs['end_id'] = data[-1]['id']
                kwargs['page'] += 1
                data = self.get_wish_history(authkey_url, **kwargs)
                if data is not None:
                    full_wish += data
        return full_wish if len(full_wish) > 0 else None

    def add_new_wishes_to_local(self,  old_wishes, new_wishes: list=[]):
        wishes = []
        for new in new_wishes:
            if new not in old_wishes:
                wishes.append(new)
        return wishes + old_wishes

    def get_local_wishhistory(self, uid: int, new_wishes: list= []):
        '''
        updates the local raw history with new ones
        '''

        old_wishes = []
        if exists(self.wish_rawfolder+"/"+str(uid)+".json"):
            with open(self.wish_rawfolder+"/"+str(uid)+".json", 'r') as f:
                old_wishes = load(f)
        wishes = self.add_new_wishes_to_local(old_wishes, new_wishes)

        return wishes
            
    def save_wishes(self, authkey: str, uid: int=None):
        if 'paimon.moe' in authkey or authkey.split('.')[-1] in ['xlsx','xls']:
            if uid is not None:
                new_wishes = self.import_from_paimonmoe(authkey, uid)
            else:
                return None
        else:
            new_wishes = self.get_full_wishhistory(authkey)
            uid = new_wishes[0]['uid']
        if new_wishes is not None:
            all_wishes = self.get_local_wishhistory(uid, new_wishes)

            with open(self.wish_rawfolder+"/"+str(uid)+".json", 'w') as f:
                dump(all_wishes, f, indent=1)
            self.process_wishes(uid)
            return True

    
    def get_region(self, uid: int):
        '''

        gets server region from uid

        '''
        uid = str(uid)[0]
        if uid == '8':
            return 'asia'
        if uid == '7':
            return 'eu'
        if uid == '6':
            return 'na'

    def get_banner(self, banner_id:str, uid: int, pull_time:str):
        '''

        gets the banner dict depening on the type of banner and pull time


        banner_id: 301, 302, 200, 100
        uid: got from the request made
        pull_time: every pull has a time in json received from api request

        '''

        banner = self.BANNER_MAP[banner_id]
        region = self.get_region(uid)
        for bann_ in self.banners[banner]:
            if 'reg_times' in bann_:
                if 'start' not in bann_['reg_times'][region]:
                    start = datetime(2999, 12, 30, 0, 0, 0)
                else:                
                    start = datetime.strptime(bann_['reg_times'][region]['start'],'%Y-%m-%d %H:%M:%S')      
                if 'end' not in bann_['reg_times'][region]:
                    end = datetime(2999, 12, 30, 23, 59, 59)
                else:
                    end = datetime.strptime(bann_['reg_times'][region]['end'],'%Y-%m-%d %H:%M:%S')           
                    
                
                if start < datetime.strptime(pull_time, '%Y-%m-%d %H:%M:%S') < end:
                    return bann_
                else:
                    if bann_['start'] == '':
                        start = datetime(2999, 12, 30, 0, 0, 0)
                    else:                
                        start = datetime.strptime(bann_['start'],'%Y-%m-%d %H:%M:%S')      
                    if bann_['end'] == '':
                        end = datetime(2999, 12, 30, 23, 59, 59)
                    else:
                        end = datetime.strptime(bann_['end'],'%Y-%m-%d %H:%M:%S')
                    if start < datetime.strptime(pull_time, '%Y-%m-%d %H:%M:%S') < end:
                            return bann_
            
    def get_banner_tally(self, tally_list: list, banner: dict):
        '''

        gets the banner from tally list

        returns none if it doesnot exist in tally

        '''
        
        banner_to_check_values = set([banner[v] for v in banner if v in ['title', 'start', 'end']])
        for bann in tally_list:
            checker = set([bann[v] for v in bann if v in ['title', 'start', 'end']])
            if len(checker.intersection(banner_to_check_values)) == 3:
                return tally_list[tally_list.index(bann)]
            

    def get_last_banner_pity_tally(self, tally_list: list, banner_type: str, star: str):
        '''

        gets last banner pity of same banner id

        banner_type -> BANNER_MAP[banner_id] -> novice, standard, characters, weapons
        stars: star 5 or 4

        '''
        pity = 1
        if len(tally_list) > 0:
            for banner in tally_list:
                if banner['type'] == banner_type:
                    if f'{star}_pity' in banner:
                        pity = banner[f'{star}_pity']
        logt('\n', f'Last Banner {star} pities list', pity, '\n')
        return pity
    def get_last_banner_rateup_tally(self, tally_list: list, banner_type: str):
        '''
        gets last rateup of banner
        used to transter rateup from one banner to another

        ------------------

        banner_type -> BANNER_MAP[banner_id] -> novice, standard, characters, weapons

        rateup is used to determine if person is at 50/50 or not
        its value is True if a person is at 50/50

        '''
        banns = []
        if len(tally_list) > 0:
            for banner in tally_list:
                if banner['type'] == banner_type:
                    banns.append(banner)
                    
        return banns[-1]['rateup'] if len(banns)> 0 else False

    def featured_item(self, banner: dict, item_name:str):
        '''

        returns a bool

        if a pulled item is a featured item of banner or if its not

        '''
        for itm in banner['featured']:
            if item_name.lower() in itm.lower():
                return True
        return False

    def process_wishes(self, uid: int):
        '''

        Script that does all
        counts pity for seperate banners
        and returns it in dict form
        
        '''
        wish_tally = []
        if exists(self.wish_rawfolder+"/"+str(uid)+".json"):
            with open(self.wish_rawfolder+"/"+str(uid)+".json", 'r') as f:
                wish_data = load(f)
        else:
            return None
        banner_stats = {}
        for banner in self.BANNER_MAP:
            banner_stats[self.BANNER_MAP[banner]] =  {
                '5050': 0,
                'primogems': 0,
                'pulls': 0,
                '5_pity': 0,
                '4_pity': 0,
                '5star_items': [],
                '4star_items': []
            }
        for pull in wish_data[::-1]:


            '''


            CREATING BANNER OBJECT

            '''
            fivepity = self.get_last_banner_pity_tally(wish_tally, self.BANNER_MAP[pull['gacha_type']], '5')
            fourpity = self.get_last_banner_pity_tally(wish_tally, self.BANNER_MAP[pull['gacha_type']], '4')     
            at_fifty = False
            wish_banner = None
            print(pull['name'], pull['time'])
            banner = self.get_banner(pull['gacha_type'], pull['uid'], pull['time'])
            banner_check = self.get_banner_tally(wish_tally, banner)            
            if banner_check is None:
                wish_tally.append({
                    'title': banner['title'],
                    'featured': [n['name'] for n in banner['items']],
                    'start': banner['start'],
                    'end': banner['end'],
                    'times' : banner['reg_times'],
                    '5_pity': fivepity,
                    '4_pity' : fourpity,
                    'pity': 0,
                    'pulls': [],        
                    'rateup': self.get_last_banner_rateup_tally(wish_tally, self.BANNER_MAP[pull['gacha_type']]),
                    '5050': 0,
                    'type': self.BANNER_MAP[pull['gacha_type']],
                    'img': banner['img']
                })
                wish_banner = wish_tally[-1]
            else:
                wish_banner = banner_check
            print('BANNER', {k : wish_banner[k] for k in wish_banner if k in ['title', 'start', 'end']})

       
            '''

            TRANSFERRING PITY BETWEEN BANNERS

            '''

            

            pull_pity = 0
            rateup = None
            

            if str(pull['rank_type']) == '4':
                pull_pity = fourpity                  
                banner_stats[self.BANNER_MAP[pull['gacha_type']]]['4star_items'].append(pull['name'] + '('+str(pull_pity)+')')
                fourpity = 0
            if str(pull['rank_type']) == '5':
                pull_pity = fivepity
                
                banner_stats[self.BANNER_MAP[pull['gacha_type']]]['5star_items'].append(pull['name'] + '('+str(pull_pity)+')')
                fivepity = 0
                rateup = self.get_last_banner_rateup_tally(wish_tally, self.BANNER_MAP[pull['gacha_type']])        
                
                if self.featured_item(wish_banner, pull['name']):
                    
                    if rateup == True:                     
                                           
                        wish_banner['5050'] += 1                                            
                        banner_stats[self.BANNER_MAP[pull['gacha_type']]]['5050'] += 1
                        at_fifty = True 
                    wish_banner['rateup'] = not wish_banner['rateup']
                else:                                     
                    wish_banner['rateup'] = False
                    #print('not a featured item', wish_banner['rateup'])
                
                
            

            #logt('\n\n\n','\n5 pity', fivepity,'|', '4 pity', fourpity,'\n','Banner', self.BANNER_MAP[pull['gacha_type']].title(), '|', 'CODE', pull['gacha_type'], '\n', pull['name'] ,'| RANK', pull['rank_type'], '| PITY', pull_pity,'\n','--------------------------')
            banner_stats[self.BANNER_MAP[pull['gacha_type']]]['primogems'] += 160
            banner_stats[self.BANNER_MAP[pull['gacha_type']]]['pulls'] += 1
            
            fourpity += 1
            fivepity += 1
            
            banner_stats[self.BANNER_MAP[pull['gacha_type']]]['5_pity'] = fivepity
            banner_stats[self.BANNER_MAP[pull['gacha_type']]]['4_pity'] = fourpity
            '''

            PULL OBJECT

            '''
            
                



            
            wish_banner['pulls'].append({
                'name': pull['name'],
                'rarity': pull['rank_type'],
                'pity': pull_pity,
                'at_fifty': at_fifty
            })

            if fourpity > 10:
                fourpity = 1
            if fivepity > 90 and self.BANNER_MAP[pull['gacha_type']] in ['characters', 'standard', 'novice']:
                fivepity = 1

            if fivepity > 80 and self.BANNER_MAP[pull['gacha_type']] == 'weapons':
                fivepity = 1
            #logt('\n', 'PITIES AFTER PROCESSNG A SINGLE PULL\n-----\n5-pity', fivepity, '| 4-pity',fourpity,'\n','-----')

            wish_banner['5_pity'] = fivepity
            wish_banner['4_pity'] = fourpity
            #logt('\n\n\n', 'banner', wish_banner['title'], '\n', 'PITIES','\n\n', wish_banner['5_pity'], wish_banner['4_pity'],'\n\n\n')
        with open(self.wish_processedfolder+"/"+wish_data[0]['uid']+".json", 'w') as f:
            dump({**{
                'data': wish_tally
            }, **banner_stats}, f, indent=1)

    
        
    '''
    authkeyurl = 'authkey url here'
    wishes = get_full_wishhistory(authkeyurl)

    with open('test_wish.json', 'w') as f:
        dump(wishes, f, indent=1)

    data = process_wishes(wishes)

    with open('test_tally.json', 'w') as f:
        dump(data, f, indent=1)

    '''

    def get_bannertype_from_name(self, banner_name:str):
        '''

        returns banner code from banner title

        '''
        for check in self.BANNER_SECOND_MAP:
            if check.lower() in banner_name.lower():
                return self.BANNER_SECOND_MAP[check]
        return '301'

    def import_from_paimonmoe(self, link: str, uid: int):
        '''

        import from paimon moe excel sheet

        either link or path

        uid: int

        '''
        sheet_file = None
        if 'http' in link:
            with requests.get(link) as r:
                if r.status_code == 200:
                    sheet_file = BytesIO(r.content)
        else:
            if exists(link):
                with open(link, 'r') as f:
                    sheet_file = BytesIO(f.read())

        if sheet_file is not None:
            wish_data= openpyxl.load_workbook(sheet_file)
            pulls = []
            for sheetn in wish_data.sheetnames[:4]:
                sheet = wish_data[sheetn]
                for row in sheet.iter_rows(min_row=1, values_only=True):    
                    values = row
                    name = values[1]
                    time = values[2]
                    if time != 'Time':
                        rarity = values[3]
                        banner = values[-1]
                        gacha_id = self.get_bannertype_from_name(banner)
                        pulls.append({
                            'uid' : str(uid),
                            'name': name,
                            'time': time,
                            'rank_type': rarity,
                            'gacha_type': gacha_id
                        })   
            return pulls[::-1]


    def get_wish_stats(self, uid: int, banner: str):
        '''

        get specific banner stats

        banner: characters, weapons, novice, standard

        '''
        if exists(self.wish_processedfolder+"/"+str(uid)+".json"):
            data = None
            with open(self.wish_processedfolder+"/"+str(uid)+".json", 'r') as f:
                data = load(f)

            if data is not None:
                
                stats = data[banner] if banner in data else {}

                banners = []
                for bann in data['data']:
                    banners.append({
                        'title': bann['title'],
                        'pull_count': len(bann['pulls']),
                        'img': bann['img'],
                        'primos': len(bann['pulls'])*160,
                        'start': 'N/A' if bann['start'] == '' else bann['start'],
                        'end': 'N/A' if bann['end'] == '' else bann['end'],
                        'pulls': [{
                            'name': itm['name'],
                            'pity': itm['pity'],
                            'star': itm['rarity'],
                            '5050': 'YES' if itm['at_fifty'] else 'NO'
                            
                        } for itm in bann['pulls']]
                    })
                
                return {
                    'banner_type': banner,
                    'stats': stats,
                    'banners': banners
                }
        return None
 

    def create_table(self, headings: list, items_list: list):
        tb = pt(headings)
        for c,row in enumerate(items_list):
            tb.add_row(row.split("|"))

        text = str(tb)
        return text

    def wish_embeds(self, banner_data: dict, author: Member):
        embeds = []
        if len(banner_data) != 0:
            stats = banner_data.get('stats', None)
            if stats is not None:
                desc_ = ''
                for s in stats:
                    if s not in ['5star_items', '4star_items']:
                        desc_ += f"**{s.title().replace('_',' ',99)}** : *{stats[s]}*\n"
                    else:
                        if s != '4star_items':
                            table = ','.join(stats[s])
                            desc_ += f"*{s.replace('star', ' star',1).title().replace('_',' ',99)} Overview*\n```css\n{table}```\n"

                embed = Embed(title=f"{banner_data['banner_type'].title()} Basic Stats", description=desc_, color=self.bot.resource_manager.get_color_from_image(author.avatar.url))
                embed.set_footer(text=f"{banner_data['banner_type'].title()} -  Basic Stats")
                embed.set_author(name=author.display_name, icon_url=author.display_avatar.url)   
                embed.set_thumbnail(url=author.display_avatar.url)         
                embeds.append(embed)

            banners = banner_data.get("banners", None)
            if banners is not None:
                for bann in banners:
                    
                    headings = ['pity', 'name', 'at 5050']
                    items = [f"{i['pity']}|{i['name']}|{i['5050']}" for i in bann['pulls']]
                    if len(items) > 36:
                        for p in range(len(items) // 36):

                            table = self.create_table(headings, items[p*36:p*36+ 36])
                            description = f"**Start date:** : {bann['start']}\n**End date:** {bann['end']}\n**Primogems used:** {bann['primos']}\n**Total Pulls:** {bann['pull_count']}\n**Items Pulled**\n```css\n{table}```"
                            embed = Embed(title=bann['title'],description=description, color=self.bot.resource_manager.get_color_from_image(bann['img']), url=bann['img'])
                            embed.set_author(name=author.display_name, icon_url=author.display_avatar.url)             
                            embeds.append(embed)
                    else:
                        table = self.create_table(headings, items)
                        description = f"**Start date:** : {bann['start']}\n**End date:** {bann['end']}\n**Primogems used:** {bann['primos']}\n**Total Pulls:** {bann['pull_count']}\n**Items Pulled**\n```css\n{table}```"
                        embed = Embed(title=bann['title'],description=description, color=self.bot.resource_manager.get_color_from_image(bann['img']), url=bann['img'])
                        embed.set_author(name=author.display_name, icon_url=author.display_avatar.url)             
                        embeds.append(embed)

        
        return embeds if len(embeds) > 0 else None

    def wishes_clear(self, uid: int):
        if exists(self.wish_processedfolder+"/"+str(uid)+".json"):
            remove(self.wish_processedfolder+"/"+str(uid)+".json")
        if exists(self.wish_rawfolder+"/"+str(uid)+".json"):
            remove(self.wish_rawfolder+"/"+str(uid)+".json")
        return True
